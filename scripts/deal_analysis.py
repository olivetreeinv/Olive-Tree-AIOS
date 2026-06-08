#!/usr/bin/env python3
"""
Olive Tree Investments — Deal Analysis Script

Modes:
  --analyze        Calculate underwriting metrics and output recommendation
  --log-deal       Append a deal row to the Deal Sourcing sheet
  --populate-analyzer  Download Deal Analyzer xlsx, write INPUTS, re-upload
  --fetch-docs     Download a doc from Google Drive for parsing

Usage examples:
  # Analyze with inline numbers
  python3 scripts/deal_analysis.py --analyze \\
    --property "Maple Terrace" --zip 30341 --asking 2800000 --units 24 \\
    --current-gpr 22000 --current-opex 11000 --repair 240000

  # Log a deal to the sheet
  python3 scripts/deal_analysis.py --log-deal \\
    --property "Maple Terrace" --address "123 Main St, Chamblee, GA" \\
    --market "Chamblee, GA" --zip 30341 --units 24 --asking 2800000 \\
    --stage "Analyzing" --broker-name "John Smith" --broker-email "j@firm.com" \\
    --platform "Crexi" --notes "PURSUE LOI — IRR 17.2%, EM 2.18x"

  # Populate Deal Analyzer with inputs
  python3 scripts/deal_analysis.py --populate-analyzer \\
    --property "Maple Terrace" --asking 2800000 --units 24 --repair 240000 \\
    --entry-cap 5.5 --exit-cap 6.0 --vintage 1978

  # Dry run — print analysis without writing anything
  python3 scripts/deal_analysis.py --analyze --dry-run \\
    --property "Test Deal" --zip 30341 --asking 1800000 --units 18 \\
    --current-gpr 17000 --current-opex 9000
"""

import argparse
import base64
import importlib
import io
import json
import sys
import os
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime

import requests
from gws_auth import get_token

# ─────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────

SPREADSHEET_ID   = "1VxOlof56s8GosrWkSctL-FMm7AoJKJ6YtM3GllMKVH4"
DEAL_ANALYZER_ID = "14bpvhKEuG4UipIDWIZC2Hud9D0JiV2X6"

SHEETS_BASE = "https://sheets.googleapis.com/v4/spreadsheets"
DRIVE_BASE  = "https://www.googleapis.com/drive/v3/files"
UPLOAD_BASE = "https://www.googleapis.com/upload/drive/v3/files"

TODAY = date.today().strftime("%m/%d/%Y")
TODAY_ISO = date.today().strftime("%Y-%m-%d")

BUY_BOX = {
    "30341": {"market": "Chamblee, GA",           "ppu_min": 90000,  "ppu_max": 140000},
    "30080": {"market": "Smyrna, GA",             "ppu_min": 110000, "ppu_max": 160000},
    "30005": {"market": "Alpharetta, GA",         "ppu_min": 140000, "ppu_max": 250000},
    "37207": {"market": "North Nashville, TN",    "ppu_min": 0,      "ppu_max": 999999},
    "37115": {"market": "Madison, TN",            "ppu_min": 0,      "ppu_max": 999999},
    "37408": {"market": "Chattanooga Southside, TN", "ppu_min": 0,   "ppu_max": 999999},
    "35801": {"market": "Huntsville Core, AL",    "ppu_min": 0,      "ppu_max": 999999},
    "35205": {"market": "Birmingham Urban, AL",   "ppu_min": 0,      "ppu_max": 999999},
    "35806": {"market": "Huntsville Growth, AL",  "ppu_min": 0,      "ppu_max": 999999},
}

# Hard thresholds
THRESHOLDS = {
    "coc_yr3":        0.08,   # 8%+ by year 3-4
    "irr":            0.15,   # 15%+
    "equity_multiple": 2.09,
    "dscr":           1.20,
    "rule_75_pct":    0.75,   # all-in ≤ 75% of stabilized value
}

# T-12 parsing keywords — shared by openpyxl and pandas parsers
_T12_INCOME_KEYS  = ["gross potential rent", "gross potential income", "scheduled base rent",
                     "gross rents", "rental income", "total income", "total revenue",
                     "effective gross income"]
_T12_EXPENSE_KEYS = ["total expenses", "total operating expenses", "operating expenses", "total opex"]
_T12_NOI_KEYS     = ["net operating income", "noi"]
_T12_VACANCY_KEYS = ["vacancy", "vacancy loss", "credit loss"]

# Rent roll parsing keywords — shared by openpyxl and pandas parsers
_RR_RENT_COLS   = ["rent", "current rent", "actual rent", "monthly rent", "contract rent", "lease rent"]
_RR_MARKET_COLS = ["market rent", "market", "asking rent"]
_RR_BED_COLS    = ["bed", "br", "bedroom", "unit type", "type", "floorplan"]
_RR_UNIT_COLS   = ["unit", "apt", "apartment", "unit no", "unit #", "unit#"]
_RR_ANCHOR_COLS = _RR_UNIT_COLS + _RR_BED_COLS + _RR_RENT_COLS + ["tenant", "status", "occupied"]

# Excel type detection signals — shared by openpyxl and pandas detectors
_EXCEL_RR_SIGNALS  = ["tenant", "unit no", "apt", "lease", "move-in"]
_EXCEL_T12_SIGNALS = ["operating expenses", "net operating", "vacancy loss", "gross potential"]

# Deal Sourcing tab column order (must match sheets_update.py rebuild)
DEAL_SOURCING_HEADERS = [
    "Market", "Zip Code", "Property Name", "Address", "Doors",
    "Asking Price", "Offer Price", "Price/Unit", "Vintage", "Cap Rate",
    "Gross Rent", "NOI", "Platform", "Brokerage", "Broker Name",
    "Broker Email", "Broker Phone", "Stage", "Date Found", "Last Updated", "Notes"
]

# ─────────────────────────────────────────────
# Auth
# ─────────────────────────────────────────────

def auth(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ─────────────────────────────────────────────
# Underwriting math — aligned with Deal Analyzer spreadsheet
# ─────────────────────────────────────────────

def _compute_irr(cash_flows, tol=1e-7, max_iter=200):
    """Newton-Raphson IRR on annual cash flows. Returns None if no convergence."""
    if len(cash_flows) < 2 or cash_flows[0] >= 0:
        return None
    for guess in (0.10, 0.05, 0.20, 0.30, 0.01, -0.05):
        r = guess
        try:
            for _ in range(max_iter):
                npv  = sum(cf / (1 + r) ** t for t, cf in enumerate(cash_flows))
                dnpv = -sum(t * cf / (1 + r) ** (t + 1) for t, cf in enumerate(cash_flows) if t)
                if abs(dnpv) < 1e-12:
                    break
                r2 = r - npv / dnpv
                if abs(r2 - r) < tol:
                    if -0.99 < r2 < 10.0:
                        return r2
                    break
                r = r2
        except (OverflowError, ZeroDivisionError, ValueError):
            continue
    return None


def _pi_payment(loan, rate_annual, amort_years):
    """Monthly P+I payment on a fully-amortizing loan."""
    r = rate_annual / 12
    n = amort_years * 12
    if r == 0 or n == 0:
        return loan / max(n, 1)
    return loan * r * (1 + r) ** n / ((1 + r) ** n - 1)


def _loan_balance(loan, rate_annual, amort_years, months_paid):
    """Remaining balance after `months_paid` P+I payments."""
    r  = rate_annual / 12
    n  = amort_years * 12
    pmt = _pi_payment(loan, rate_annual, amort_years)
    if r == 0:
        return max(0.0, loan - pmt * months_paid)
    return max(0.0, loan * (1 + r) ** months_paid - pmt * ((1 + r) ** months_paid - 1) / r)


def calculate_metrics(args):
    """
    Core underwriting aligned with the Olive Tree Deal Analyzer spreadsheet.

    Required: asking, units
    Key defaults (match spreadsheet): ltv=0.70, bridge_rate=0.0675, hold_years=6,
      vacancy_pct=0.10, closing_costs_pct=0.06, io_years=2, amort_years=25
    """
    # ── Inputs ──
    asking        = args.asking
    offer         = getattr(args, 'offer', None) or asking
    units         = args.units
    repair        = getattr(args, 'repair', 0) or 0

    ltv           = getattr(args, 'ltv', 0.70)               # 70% (spreadsheet)
    closing_pct   = getattr(args, 'closing_costs_pct', 0.06) # 6% (spreadsheet)
    interest_rate = getattr(args, 'bridge_rate', 0.0675)     # 6.75% (spreadsheet)
    io_years      = getattr(args, 'io_years', 2)             # 2yr I/O (spreadsheet)
    amort_years   = getattr(args, 'amort_years', 25)         # 25yr am. (spreadsheet)
    hold_years    = getattr(args, 'hold_years', 6)           # 6yr hold (spreadsheet)
    vacancy_pct   = getattr(args, 'vacancy_pct', 0.10)       # 10% (spreadsheet)
    rent_growth   = getattr(args, 'rent_growth', 0.03)       # 3%/yr from yr3
    expense_growth = getattr(args, 'expense_growth', 0.02)   # 2%/yr
    other_income_annual = (getattr(args, 'other_income', 0) or 0)
    selling_costs_pct = 0.05

    current_gpr_mo  = getattr(args, 'current_gpr', None)   # monthly
    market_gpr_mo   = getattr(args, 'market_gpr', None)    # monthly (proforma)
    current_opex_mo = getattr(args, 'current_opex', None)  # monthly
    entry_cap       = getattr(args, 'entry_cap', None)
    exit_cap        = getattr(args, 'exit_cap', None)

    # ── Sources & uses ──
    loan_amount   = offer * ltv
    equity_invest = offer * (1 - ltv) + repair + offer * closing_pct
    all_in        = offer + repair + offer * closing_pct

    # ── Debt service ──
    io_pmt_annual  = loan_amount * interest_rate
    pi_pmt_annual  = _pi_payment(loan_amount, interest_rate, amort_years) * 12

    # ── Annualised rents ──
    current_gpr_annual = (current_gpr_mo * 12) if current_gpr_mo else None
    market_gpr_annual  = (market_gpr_mo * 12)  if market_gpr_mo  else current_gpr_annual
    current_opex_annual = (current_opex_mo * 12) if current_opex_mo else None

    # Spreadsheet: yr1 = current rents, yr2 steps up to proforma rents
    yr1_gpr = current_gpr_annual or market_gpr_annual or 0
    yr2_gpr = market_gpr_annual  or current_gpr_annual or 0

    # Proforma opex baseline (yr1); grows 2%/yr after
    if current_opex_annual:
        proforma_opex = current_opex_annual
    elif yr2_gpr:
        proforma_opex = yr2_gpr * (1 - vacancy_pct) * 0.42  # 42% of EGI (spreadsheet avg)
    else:
        proforma_opex = 0

    # ── Current NOI & entry cap ──
    current_noi_annual = None
    if yr1_gpr and current_opex_annual:
        current_egi = yr1_gpr * (1 - vacancy_pct) + other_income_annual
        current_noi_annual = current_egi - current_opex_annual
    if entry_cap is None and current_noi_annual:
        entry_cap = (current_noi_annual / offer) * 100

    # ── Multi-year NOI & cash flows (spreadsheet PROFORMA logic) ──
    annual_noi, annual_cf, annual_ds = [], [], []

    for yr in range(1, hold_years + 1):
        # Rents: yr1=current, yr2=proforma, yr3+=proforma×(1+g)^(yr-2)
        if yr == 1:
            gpr = yr1_gpr
        elif yr == 2:
            gpr = yr2_gpr
        else:
            gpr = yr2_gpr * (1 + rent_growth) ** (yr - 2)

        egi  = gpr * (1 - vacancy_pct) + other_income_annual
        opex = proforma_opex * (1 + expense_growth) ** (yr - 1)
        noi  = egi - opex
        annual_noi.append(noi)

        ds = io_pmt_annual if yr <= io_years else pi_pmt_annual
        annual_ds.append(ds)
        annual_cf.append(noi - ds)

    # ── Exit / sale ──
    if exit_cap is None and entry_cap is not None:
        exit_cap = entry_cap + 0.5   # conservative cap expansion

    sale_price = net_sale_proceeds = loan_at_sale = None
    if exit_cap and exit_cap > 0 and annual_noi:
        sale_price   = annual_noi[-1] / (exit_cap / 100)
        pi_months    = max(0, hold_years - io_years) * 12
        loan_at_sale = _loan_balance(loan_amount, interest_rate, amort_years, pi_months)
        net_sale_proceeds = sale_price * (1 - selling_costs_pct) - loan_at_sale

    # ── Levered cash flows → IRR ──
    levered_cfs = [-equity_invest] + annual_cf[:]
    if net_sale_proceeds is not None:
        levered_cfs[-1] += net_sale_proceeds

    irr_estimate = _compute_irr(levered_cfs)

    # ── Returns ──
    coc_yr1 = annual_cf[0] / equity_invest if equity_invest and annual_cf else None
    coc_yr3 = annual_cf[2] / equity_invest if equity_invest and len(annual_cf) >= 3 else None

    equity_multiple = None
    if equity_invest and net_sale_proceeds is not None:
        total_return = sum(annual_cf) + net_sale_proceeds
        equity_multiple = total_return / equity_invest

    # ── DSCR (year 1) ──
    dscr = annual_noi[0] / annual_ds[0] if annual_noi and annual_ds[0] else None

    # ── Stabilized value & 75% rule (uses yr1 proforma NOI) ──
    stabilized_noi   = annual_noi[0] if annual_noi else None
    stabilized_value = None
    rule_75_ratio = rule_75_pass = None
    if stabilized_noi and exit_cap:
        stabilized_value = stabilized_noi / (exit_cap / 100)
        rule_75_ratio = all_in / stabilized_value
        rule_75_pass  = rule_75_ratio < THRESHOLDS["rule_75_pct"]

    # ── 1% rule ──
    ppu = offer / units
    avg_rent_mo = (current_gpr_mo / units) if current_gpr_mo else None
    rule_1pct_ratio = (avg_rent_mo / ppu) if avg_rent_mo else None
    rule_1pct_pass  = (rule_1pct_ratio >= 0.01) if rule_1pct_ratio is not None else None

    # ── 10x NOI rule ──
    rule_10x_noi  = offer / current_noi_annual if current_noi_annual else None
    rule_10x_pass = rule_10x_noi <= 10 if rule_10x_noi is not None else None

    # ── PPU vs buy box ──
    zip_str = str(getattr(args, 'zip', '') or '')
    ppu_in_range = None
    if zip_str in BUY_BOX:
        bb = BUY_BOX[zip_str]
        ppu_in_range = True if bb["ppu_min"] == 0 else (bb["ppu_min"] <= ppu <= bb["ppu_max"])

    return {
        "asking":            asking,
        "offer":             offer,
        "units":             units,
        "ppu":               ppu,
        "repair":            repair,
        "all_in":            all_in,
        "loan_amount":       loan_amount,
        "equity_invested":   equity_invest,
        "annual_debt_svc":   io_pmt_annual,   # legacy key
        "io_pmt_annual":     io_pmt_annual,
        "pi_pmt_annual":     pi_pmt_annual,
        "current_noi":       current_noi_annual,
        "stabilized_noi":    stabilized_noi,
        "entry_cap":         entry_cap,
        "exit_cap":          exit_cap,
        "stabilized_value":  stabilized_value,
        "sale_price":        sale_price,
        "loan_at_sale":      loan_at_sale,
        "net_sale_proceeds": net_sale_proceeds,
        "annual_noi":        annual_noi,
        "annual_cf":         annual_cf,
        "annual_ds":         annual_ds,
        "levered_cfs":       levered_cfs,
        "dscr":              dscr,
        "coc_yr1":           coc_yr1,
        "coc_yr3":           coc_yr3,
        "equity_multiple":   equity_multiple,
        "irr_estimate":      irr_estimate,
        "rule_75_ratio":     rule_75_ratio,
        "rule_75_pass":      rule_75_pass,
        "rule_1pct_ratio":   rule_1pct_ratio,
        "rule_1pct_pass":    rule_1pct_pass,
        "rule_10x_noi":      rule_10x_noi,
        "rule_10x_pass":     rule_10x_pass,
        "ppu_in_range":      ppu_in_range,
    }


def score_deal(metrics, zip_str):
    """
    Returns (recommendation, passes, fails, warnings).
    recommendation: 'PURSUE_LOI' | 'MORE_INFO' | 'PASS'
    """
    passes   = []
    fails    = []
    warnings = []

    def chk(label, value, threshold, direction="ge"):
        if value is None:
            warnings.append(f"{label}: N/A (missing data)")
            return
        ok = (value >= threshold) if direction == "ge" else (value <= threshold)
        fv = fmt_pct(value)     if value     < 1 else fmt_num(value,     "")
        ft = fmt_pct(threshold) if threshold < 1 else fmt_num(threshold, "")
        entry = f"{label}: {fv} (threshold: {ft})"
        (passes if ok else fails).append(entry)

    if zip_str and zip_str not in BUY_BOX:
        fails.append(f"ZIP {zip_str} is outside the active buy box")

    chk("DSCR",            metrics["dscr"],            THRESHOLDS["dscr"])
    chk("Cash-on-Cash Yr3", metrics["coc_yr3"],        THRESHOLDS["coc_yr3"])
    chk("IRR (est.)",      metrics["irr_estimate"],    THRESHOLDS["irr"])
    chk("Equity Multiple", metrics["equity_multiple"], THRESHOLDS["equity_multiple"])

    if metrics["rule_75_pass"] is not None:
        entry = f"75% Rule: {metrics['rule_75_ratio']:.1%} all-in/stabilized (threshold: <75%)"
        (passes if metrics["rule_75_pass"] else fails).append(entry)
    else:
        warnings.append("75% Rule: N/A (need stabilized value estimate)")

    if metrics["rule_1pct_pass"] is not None:
        entry = f"1% Rule: {metrics['rule_1pct_ratio']:.3%} rent/PPU (threshold: ≥1%)"
        (passes if metrics["rule_1pct_pass"] else fails).append(entry)
    else:
        warnings.append("1% Rule: N/A (need rent data)")

    if metrics.get("rule_10x_pass") is not None:
        entry = f"10x NOI Rule: {metrics['rule_10x_noi']:.1f}x offer/NOI (threshold: ≤10x)"
        (passes if metrics["rule_10x_pass"] else warnings).append(entry)

    if metrics["ppu_in_range"] is not None and not metrics["ppu_in_range"]:
        warnings.append(f"Price/unit ${metrics['ppu']:,.0f} outside buy box range for this market")

    hard_fails = len(fails)
    if hard_fails == 0:
        rec = "PURSUE_LOI"
    elif hard_fails <= 2 and len(warnings) <= 2:
        rec = "MORE_INFO"
    else:
        rec = "PASS"

    return rec, passes, fails, warnings


def fmt_pct(v):
    return f"{v:.1%}" if v is not None else "N/A"

def fmt_num(v, prefix="$"):
    if v is None:
        return "N/A"
    if isinstance(v, float) and v < 100:
        return f"{v:.2f}x"
    return f"{prefix}{v:,.0f}"


def print_analysis(args, metrics, rec, passes, fails, warnings):
    prop  = getattr(args, 'property', 'Unknown')
    addr  = getattr(args, 'address', '')
    zip_s = str(getattr(args, 'zip', '') or '')
    mkt   = BUY_BOX.get(zip_s, {}).get("market", zip_s) if zip_s else ""

    rec_label = {"PURSUE_LOI": "PURSUE LOI", "MORE_INFO": "MORE INFO NEEDED", "PASS": "PASS"}[rec]
    rec_emoji = {"PURSUE_LOI": "✅", "MORE_INFO": "⚠️", "PASS": "❌"}[rec]

    print(f"\n{'='*60}")
    print(f"  DEAL ANALYSIS — {prop.upper()}")
    print(f"{'='*60}")
    if addr:
        print(f"  Address  : {addr}")
    if mkt:
        print(f"  Market   : {mkt} ({zip_s})")
    print(f"  Asking   : ${metrics['asking']:,.0f}  |  Units: {metrics['units']}  |  PPU: ${metrics['ppu']:,.0f}")
    if metrics["repair"]:
        print(f"  Repair   : ${metrics['repair']:,.0f}  |  All-in: ${metrics['all_in']:,.0f}")
    print()

    print("  FINANCIALS")
    print(f"  {'Metric':<22} {'Current':>12} {'Stabilized':>12} {'Threshold':>12} {'Status':>8}")
    print(f"  {'-'*22} {'-'*12} {'-'*12} {'-'*12} {'-'*8}")

    def row(label, curr, stable, thresh, status):
        print(f"  {label:<22} {curr:>12} {stable:>12} {thresh:>12} {status:>8}")

    row("NOI (annual)",
        fmt_num(metrics["current_noi"]), fmt_num(metrics["stabilized_noi"]), "—", "—")
    row("Entry Cap",
        f"{metrics['entry_cap']:.2f}%" if metrics["entry_cap"] else "N/A", "—", "—", "—")
    row("Exit Cap (est.)",
        "—", f"{metrics['exit_cap']:.2f}%" if metrics["exit_cap"] else "N/A", "—", "—")
    row("DSCR",
        fmt_num(metrics["dscr"], ""), "—", "> 1.20",
        "✅" if metrics["dscr"] and metrics["dscr"] >= 1.20 else "❌" if metrics["dscr"] else "—")
    row("Cash-on-Cash Yr1",
        fmt_pct(metrics["coc_yr1"]), "—", "—", "—")
    row("Cash-on-Cash Yr3",
        "—", fmt_pct(metrics["coc_yr3"]), "≥ 8.0%",
        "✅" if metrics["coc_yr3"] and metrics["coc_yr3"] >= 0.08 else "❌" if metrics["coc_yr3"] else "—")
    row("IRR (est.)",
        "—", fmt_pct(metrics["irr_estimate"]), "≥ 15.0%",
        "✅" if metrics["irr_estimate"] and metrics["irr_estimate"] >= 0.15 else "❌" if metrics["irr_estimate"] else "—")
    row("Equity Multiple",
        "—", fmt_num(metrics["equity_multiple"], ""), "≥ 2.09x",
        "✅" if metrics["equity_multiple"] and metrics["equity_multiple"] >= 2.09 else "❌" if metrics["equity_multiple"] else "—")
    row("75% Rule",
        "—", f"{metrics['rule_75_ratio']:.1%}" if metrics["rule_75_ratio"] else "N/A",
        "< 75.0%",
        "✅" if metrics["rule_75_pass"] else "❌" if metrics["rule_75_pass"] is not None else "—")
    row("1% Rule (rent/PPU)",
        f"{metrics['rule_1pct_ratio']:.3%}" if metrics["rule_1pct_ratio"] else "N/A",
        "—", "≥ 1.000%",
        "✅" if metrics["rule_1pct_pass"] else "❌" if metrics["rule_1pct_pass"] is not None else "—")
    row("10x NOI Rule",
        f"{metrics['rule_10x_noi']:.1f}x" if metrics["rule_10x_noi"] else "N/A",
        "—", "≤ 10.0x",
        "✅" if metrics["rule_10x_pass"] else "❌" if metrics["rule_10x_pass"] is not None else "—")

    # ── Multi-year proforma ──
    annual_noi = metrics.get("annual_noi", [])
    annual_cf  = metrics.get("annual_cf",  [])
    annual_ds  = metrics.get("annual_ds",  [])
    eq_inv     = metrics.get("equity_invested") or 0
    if annual_noi:
        print()
        hold_yrs = len(annual_noi)
        print(f"  MULTI-YEAR PROFORMA ({hold_yrs}-yr hold)")
        print(f"  {'Yr':<4} {'NOI':>10} {'Debt Svc':>10} {'Cash Flow':>10} {'CoC':>7}")
        print(f"  {'─'*4} {'─'*10} {'─'*10} {'─'*10} {'─'*7}")
        for i, (noi, ds, cf) in enumerate(zip(annual_noi, annual_ds, annual_cf), 1):
            coc_s = f"{cf/eq_inv:.1%}" if eq_inv else "—"
            tag   = "  ← current rents" if i == 1 else "  ← proforma rents" if i == 2 else ""
            print(f"  {i:<4} {fmt_num(noi):>10} {fmt_num(ds):>10} {fmt_num(cf):>10} {coc_s:>7}{tag}")

    # ── Exit analysis ──
    if metrics.get("sale_price") is not None:
        ec = metrics.get("exit_cap") or 0
        hold_yrs = len(annual_noi) if annual_noi else "?"
        print()
        print(f"  EXIT ANALYSIS (yr{hold_yrs} sale @ {ec:.1f}% cap)")
        print(f"  {'─'*40}")
        print(f"  Sale Price          : {fmt_num(metrics['sale_price'])}")
        print(f"  Loan Balance        : {fmt_num(metrics['loan_at_sale'])}")
        print(f"  Net Sale Proceeds   : {fmt_num(metrics['net_sale_proceeds'])}")
        if metrics.get("equity_multiple"):
            print(f"  Equity Multiple     : {metrics['equity_multiple']:.2f}x")
        if metrics.get("irr_estimate"):
            print(f"  Levered IRR         : {metrics['irr_estimate']:.1%}")

    print()
    if passes:
        print("  PASSES:")
        for p in passes:
            print(f"    ✅ {p}")
    if fails:
        print("  FAILS:")
        for f in fails:
            print(f"    ❌ {f}")
    if warnings:
        print("  WARNINGS / MISSING DATA:")
        for w in warnings:
            print(f"    ⚠️  {w}")

    print()
    print(f"  {'─'*56}")
    print(f"  RECOMMENDATION: {rec_emoji} {rec_label}")
    print(f"  {'─'*56}")

    if rec == "PURSUE_LOI":
        print("  → Ready to draft LOI. Reply 'draft LOI' or run /lets-get-to-work.")
    elif rec == "MORE_INFO":
        print("  → Missing data may change this. Request docs before deciding.")
    else:
        print("  → Deal doesn't fit thresholds. Log as Pass in Deal Sourcing.")

    print(f"{'='*60}\n")

    return rec


# ─────────────────────────────────────────────
# Sheets — log deal
# ─────────────────────────────────────────────

def log_deal(token, args, metrics=None):
    prop     = getattr(args, 'property', '')
    address  = getattr(args, 'address', '')
    market   = getattr(args, 'market', '')
    zip_s    = str(getattr(args, 'zip', '') or '')
    units    = getattr(args, 'units', '')
    asking   = getattr(args, 'asking', '')
    stage    = getattr(args, 'stage', 'New')
    platform = getattr(args, 'platform', '')
    brokerage   = getattr(args, 'brokerage', '')
    broker_name = getattr(args, 'broker_name', '')
    broker_email= getattr(args, 'broker_email', '')
    broker_phone= getattr(args, 'broker_phone', '')
    vintage  = getattr(args, 'vintage', '')
    notes    = getattr(args, 'notes', '')

    # infer from metrics if available
    offer    = ''
    cap_rate = ''
    gross_rent = ''
    noi_val  = ''
    if metrics:
        cap_rate   = f"{metrics['entry_cap']:.2f}%" if metrics.get('entry_cap') else ''
        gross_rent = f"${metrics['current_noi']:,.0f}" if metrics.get('current_noi') else ''
        noi_val    = f"${metrics['stabilized_noi']:,.0f}" if metrics.get('stabilized_noi') else ''

    if not market and zip_s in BUY_BOX:
        market = BUY_BOX[zip_s]["market"]

    # ppu formula placeholder — sheet formula will auto-calc from Asking/Doors columns
    row = [
        market, zip_s, prop, address, units,
        asking, offer, '',       # Price/Unit = formula in sheet
        vintage, cap_rate, gross_rent, noi_val,
        platform, brokerage, broker_name, broker_email, broker_phone,
        stage, TODAY, TODAY, notes
    ]

    r = requests.post(
        f"{SHEETS_BASE}/{SPREADSHEET_ID}/values/Deal%20Sourcing!A:U:append",
        headers=auth(token),
        params={"valueInputOption": "USER_ENTERED", "insertDataOption": "INSERT_ROWS"},
        json={"values": [row]},
        timeout=30,
    )
    r.raise_for_status()
    result = r.json()
    updated = result.get("updates", {}).get("updatedRange", "Deal Sourcing tab")
    print(f"✅ Logged to sheet: {updated}")
    return result


# ─────────────────────────────────────────────
# Sheets — dedup check
# ─────────────────────────────────────────────

def deal_exists(token, prop, address):
    r = requests.get(
        f"{SHEETS_BASE}/{SPREADSHEET_ID}/values/Deal%20Sourcing!A:D",
        headers=auth(token),
        timeout=30,
    )
    r.raise_for_status()
    rows = r.json().get("values", [])
    prop_lower    = prop.lower().strip()
    address_lower = address.lower().strip()
    for row in rows[1:]:
        existing_name = row[2].lower().strip() if len(row) > 2 else ""
        existing_addr = row[3].lower().strip() if len(row) > 3 else ""
        if existing_name == prop_lower or (address_lower and existing_addr == address_lower):
            return True
    return False


# ─────────────────────────────────────────────
# Drive — fetch doc
# ─────────────────────────────────────────────

def fetch_doc(token, drive_id, dest_path=None):
    if dest_path is None:
        dest_path = f"/tmp/deal_doc_{drive_id}.xlsx"
    with requests.get(
        f"{DRIVE_BASE}/{drive_id}",
        headers={"Authorization": f"Bearer {token}"},
        params={"alt": "media"},
        stream=True,
        timeout=60,
    ) as r:
        r.raise_for_status()
        with open(dest_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
    print(f"✅ Downloaded {drive_id} → {dest_path}")
    return dest_path


# ─────────────────────────────────────────────
# Gmail — fetch attachments
# ─────────────────────────────────────────────

GMAIL_BASE = "https://gmail.googleapis.com/gmail/v1/users/me"

EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
}
PDF_MIMES = {"application/pdf"}

_T12_KEYWORDS    = ["t-12", "t12", "income", "p&l", "trailing"]
_RR_KEYWORDS     = ["rent roll", "rentroll", "rent_roll", "rental", "unit mix", "unitmix"]
_OM_KEYWORDS     = ["om", "offering memorandum", "offering_memorandum", "offeringmemorandum", "memo"]


def _classify_attachment(fname):
    """Return 't12', 'rent-roll', 'om', or None based on filename."""
    name = fname.lower()
    ext  = name.rsplit(".", 1)[-1] if "." in name else ""
    if any(k in name for k in _T12_KEYWORDS) and ext in ("xlsx", "xls"):
        return "t12"
    if any(k in name for k in _RR_KEYWORDS):
        return "rent-roll"
    if any(k in name for k in _OM_KEYWORDS) and ext == "pdf":
        return "om"
    if ext in ("xlsx", "xls"):
        return "excel-unknown"  # caller will auto-detect type
    if ext == "pdf":
        return "pdf-unknown"    # caller will try OM first
    return None


def fetch_gmail_attachments(token, message_id, dest_dir="/tmp"):
    """Download xlsx/xls/pdf deal attachments from a Gmail message.

    Returns dict with keys: 't12', 'rent-roll', 'om', 'other'
    Each key maps to a local file path (str) or None / list for 'other'.
    """
    hdrs = {"Authorization": f"Bearer {token}"}

    msg = requests.get(f"{GMAIL_BASE}/messages/{message_id}?format=full", headers=hdrs, timeout=30)
    msg.raise_for_status()
    payload = msg.json().get("payload", {})

    def collect_parts(part):
        parts = []
        if part.get("filename") and part.get("body", {}).get("attachmentId"):
            parts.append(part)
        for sub in part.get("parts", []):
            parts.extend(collect_parts(sub))
        return parts

    attachment_parts = collect_parts(payload)
    if not attachment_parts:
        print("No attachments found in message.")
        return {"t12": None, "rent-roll": None, "om": None, "other": []}

    result = {"t12": None, "rent-roll": None, "om": None, "other": []}

    # Keep only downloadable deal attachments (xlsx/xls/pdf)
    wanted = []
    for part in attachment_parts:
        fname = part["filename"]
        ext   = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
        mime  = part.get("mimeType", "")
        if ext not in ("xlsx", "xls", "pdf") and mime not in (EXCEL_MIMES | PDF_MIMES):
            print(f"  Skipping: {fname}")
            continue
        wanted.append(part)

    def download(part):
        # Sanitize the filename — never trust an email-supplied path (traversal/overwrite)
        fname  = os.path.basename(part["filename"]) or "attachment"
        att_id = part["body"]["attachmentId"]
        att = requests.get(
            f"{GMAIL_BASE}/messages/{message_id}/attachments/{att_id}",
            headers=hdrs, timeout=60,
        )
        att.raise_for_status()
        data = base64.urlsafe_b64decode(att.json()["data"])
        dest = os.path.join(dest_dir, fname)
        with open(dest, "wb") as f:
            f.write(data)
        print(f"✅ Saved → {dest}")
        return fname, dest

    # Download independent attachments in parallel
    with ThreadPoolExecutor(max_workers=5) as executor:
        downloaded = list(executor.map(download, wanted))

    for fname, dest in downloaded:
        doc_type = _classify_attachment(fname)
        if doc_type in ("t12", "rent-roll", "om"):
            result[doc_type] = dest
        elif doc_type == "excel-unknown":
            result["t12"] = result["t12"] or dest   # assign to T-12 slot if open
        elif doc_type == "pdf-unknown":
            result["om"]  = result["om"]  or dest   # assign to OM slot if open
        else:
            result["other"].append(dest)

    return result


# ─────────────────────────────────────────────
# Excel parsing — rent roll & T-12
# ─────────────────────────────────────────────

def _require(pkg, install_hint=None):
    try:
        return importlib.import_module(pkg)
    except ImportError:
        hint = install_hint or f"pip3 install {pkg}"
        print(f"ERROR: {pkg} not installed. Run: {hint}")
        sys.exit(1)


def _require_pandas():
    return _require("pandas", "python3 -m pip install pandas openpyxl xlrd")


def _require_openpyxl():
    return _require("openpyxl", "pip3 install openpyxl")


def detect_excel_type_openpyxl(path):
    """Detect 'rent-roll' or 't12' by scanning first 30 rows with openpyxl."""
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        parts = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 30:
                break
            parts.extend(str(v).lower() for v in row if v is not None)
    finally:
        wb.close()
    text = " ".join(parts)
    rr  = sum(w in text for w in _EXCEL_RR_SIGNALS)
    t12 = sum(w in text for w in _EXCEL_T12_SIGNALS)
    return "rent-roll" if rr >= t12 else "t12"


def parse_t12_openpyxl(path):
    """
    Parse a T-12 / income statement xlsx with openpyxl (no pandas required).
    Returns dict: current_gpr, current_opex, current_noi (annual).
    """
    openpyxl = _require_openpyxl()

    result = {"current_gpr": None, "current_opex": None, "current_noi": None,
              "vacancy_annual": None, "source": "T-12"}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                label = str(row[0]).lower().strip()
                nums = []
                for v in row[1:]:
                    try:
                        f = float(v)
                        if f > 0:
                            nums.append(f)
                    except (TypeError, ValueError):
                        pass
                if not nums:
                    continue
                median_val = sorted(nums)[len(nums) // 2]
                annual = nums[-1] if (len(nums) > 1 and nums[-1] >= median_val * 8) else sum(nums[:12])

                if any(label.startswith(k) for k in _T12_INCOME_KEYS) and result["current_gpr"] is None:
                    result["current_gpr"] = annual
                elif any(label.startswith(k) for k in _T12_EXPENSE_KEYS) and result["current_opex"] is None:
                    result["current_opex"] = annual
                elif any(label.startswith(k) for k in _T12_NOI_KEYS) and result["current_noi"] is None:
                    result["current_noi"] = annual
                elif any(label.startswith(k) for k in _T12_VACANCY_KEYS) and result["vacancy_annual"] is None:
                    result["vacancy_annual"] = annual

            if result["current_gpr"] or result["current_noi"]:
                break
    finally:
        wb.close()

    if result["current_noi"] and result["current_gpr"] and result["current_opex"] is None:
        result["current_opex"] = result["current_gpr"] - result["current_noi"]
    if result["current_noi"] and result["current_opex"] and result["current_gpr"] is None:
        result["current_gpr"] = result["current_noi"] + result["current_opex"]

    return result


def parse_rent_roll_openpyxl(path):
    """
    Parse a rent roll xlsx with openpyxl (no pandas required).
    Returns dict: units, current_gpr (monthly total), unit_mix list.
    """
    openpyxl = _require_openpyxl()

    result = {"units": 0, "current_gpr": 0.0, "unit_mix": [], "source": "Rent Roll"}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
      for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_iter = ws.iter_rows(values_only=True)

        headers = None
        for row in row_iter:
            row_text = [str(v).lower().strip() if v is not None else "" for v in row]
            if sum(1 for v in row_text if any(k in v for k in _RR_ANCHOR_COLS)) >= 2:
                headers = row_text
                break

        if headers is None:
            continue

        rent_col   = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_RENT_COLS)), None)
        market_col = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_MARKET_COLS)), None)
        bed_col    = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_BED_COLS)), None)

        if rent_col is None:
            continue

        unit_rents, unit_markets = {}, {}
        total_gpr, unit_count = 0.0, 0

        for row in row_iter:
            if not row or len(row) <= rent_col or row[rent_col] is None:
                continue
            try:
                rent_val = float(row[rent_col])
            except (ValueError, TypeError):
                continue
            if rent_val <= 0:
                continue

            total_gpr += rent_val
            unit_count += 1

            if bed_col is not None and len(row) > bed_col and row[bed_col] is not None:
                bed = str(row[bed_col]).strip()
                unit_rents.setdefault(bed, []).append(rent_val)
                if market_col is not None and len(row) > market_col and row[market_col] is not None:
                    try:
                        mkt = float(row[market_col])
                        if mkt > 0:
                            unit_markets.setdefault(bed, []).append(mkt)
                    except (ValueError, TypeError):
                        pass

        if unit_count == 0:
            continue

        result["units"] = unit_count
        result["current_gpr"] = total_gpr
        for bed, rents in unit_rents.items():
            entry = {"type": bed, "count": len(rents),
                     "current_rent": round(sum(rents) / len(rents), 0)}
            mkts = unit_markets.get(bed, [])
            if mkts:
                entry["market_rent"] = round(sum(mkts) / len(mkts), 0)
            result["unit_mix"].append(entry)
        break
    finally:
        wb.close()

    return result


def detect_excel_type(path):
    """Guess 'rent-roll' or 't12' by scanning column/row headers."""
    pd = _require_pandas()
    with pd.ExcelFile(path) as xl:
        df = xl.parse(xl.sheet_names[0], header=None, nrows=30)
    text = " ".join(str(v).lower() for row in df.values for v in row)
    rr  = sum(w in text for w in _EXCEL_RR_SIGNALS)
    t12 = sum(w in text for w in _EXCEL_T12_SIGNALS)
    return "rent-roll" if rr >= t12 else "t12"


def parse_t12(path):
    """
    Parse a T-12 / income statement Excel.
    Returns dict: current_gpr, current_opex, current_noi (all annual totals).
    """
    pd = _require_pandas()

    result = {"current_gpr": None, "current_opex": None, "current_noi": None,
              "vacancy_annual": None, "source": "T-12"}

    with pd.ExcelFile(path) as xl:
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, header=None)

            for _, row in df.iterrows():
                label = str(row.iloc[0]).lower().strip() if len(row) > 0 else ""

                nums = []
                for v in row.iloc[1:]:
                    try:
                        f = float(v)
                        if f > 0:
                            nums.append(f)
                    except (TypeError, ValueError):
                        pass
                if not nums:
                    continue

                median_val = sorted(nums)[len(nums) // 2]
                annual = nums[-1] if (len(nums) > 1 and nums[-1] >= median_val * 8) else sum(nums[:12])

                if any(label.startswith(k) for k in _T12_INCOME_KEYS) and result["current_gpr"] is None:
                    result["current_gpr"] = annual
                elif any(label.startswith(k) for k in _T12_EXPENSE_KEYS) and result["current_opex"] is None:
                    result["current_opex"] = annual
                elif any(label.startswith(k) for k in _T12_NOI_KEYS) and result["current_noi"] is None:
                    result["current_noi"] = annual
                elif any(label.startswith(k) for k in _T12_VACANCY_KEYS) and result["vacancy_annual"] is None:
                    result["vacancy_annual"] = annual

            if result["current_gpr"] or result["current_noi"]:
                break

    # Derive missing values
    if result["current_noi"] and result["current_gpr"] and result["current_opex"] is None:
        result["current_opex"] = result["current_gpr"] - result["current_noi"]
    if result["current_noi"] and result["current_opex"] and result["current_gpr"] is None:
        result["current_gpr"] = result["current_noi"] + result["current_opex"]

    return result


def parse_rent_roll(path):
    """
    Parse a rent roll Excel.
    Returns dict: units, current_gpr (monthly total), unit_mix list.
    """
    pd = _require_pandas()

    result = {"units": 0, "current_gpr": 0.0, "unit_mix": [], "source": "Rent Roll"}

    with pd.ExcelFile(path) as xl:
        for sheet_name in xl.sheet_names:
            df_raw = xl.parse(sheet_name, header=None)

            header_row = None
            for i, row in df_raw.iterrows():
                row_text = [str(v).lower().strip() for v in row]
                hits = sum(1 for v in row_text if any(k in v for k in _RR_ANCHOR_COLS))
                if hits >= 2:
                    header_row = i
                    break
            if header_row is None:
                continue

            df = xl.parse(sheet_name, header=header_row)
            df.columns = [str(c).lower().strip() for c in df.columns]

            rent_col   = next((c for c in df.columns if any(k in c for k in _RR_RENT_COLS)), None)
            market_col = next((c for c in df.columns if any(k in c for k in _RR_MARKET_COLS)), None)
            bed_col    = next((c for c in df.columns if any(k in c for k in _RR_BED_COLS)), None)

            if rent_col is None:
                continue

            df[rent_col] = pd.to_numeric(df[rent_col], errors="coerce")
            df = df[df[rent_col] > 0].copy()

            if df.empty:
                continue

            result["units"] = len(df)
            result["current_gpr"] = float(df[rent_col].sum())

            if bed_col:
                df[bed_col] = df[bed_col].astype(str).str.strip()
                agg = {"count": (rent_col, "count"), "avg_rent": (rent_col, "mean")}
                if market_col:
                    df[market_col] = pd.to_numeric(df[market_col], errors="coerce")
                    agg["avg_market"] = (market_col, "mean")
                mix = df.groupby(bed_col).agg(**agg).reset_index()
                for _, row in mix.iterrows():
                    entry = {
                        "type": str(row[bed_col]),
                        "count": int(row["count"]),
                        "current_rent": round(float(row["avg_rent"]), 0),
                    }
                    if market_col and "avg_market" in row and not pd.isna(row["avg_market"]):
                        entry["market_rent"] = round(float(row["avg_market"]), 0)
                    result["unit_mix"].append(entry)
            break

    return result


def parse_excel_file(path, excel_type="auto"):
    """Dispatch to the right parser. Uses openpyxl by default; falls back to pandas."""
    if excel_type == "auto":
        try:
            excel_type = detect_excel_type_openpyxl(path)
        except Exception:
            excel_type = detect_excel_type(path)
    if excel_type == "rent-roll":
        try:
            return parse_rent_roll_openpyxl(path)
        except Exception:
            return parse_rent_roll(path)
    try:
        return parse_t12_openpyxl(path)
    except Exception:
        return parse_t12(path)


def print_parse_summary(parsed):
    src = parsed.get("source", "Excel")
    print(f"\n{'─'*50}")
    print(f"  PARSED: {src}")
    print(f"{'─'*50}")
    if src == "T-12":
        print(f"  GPR (annual)   : {fmt_num(parsed.get('current_gpr'))}")
        print(f"  OpEx (annual)  : {fmt_num(parsed.get('current_opex'))}")
        print(f"  NOI (annual)   : {fmt_num(parsed.get('current_noi'))}")
        if parsed.get("vacancy_annual"):
            print(f"  Vacancy (ann.) : {fmt_num(parsed.get('vacancy_annual'))}")
        # Surface as monthly for --analyze
        gpr = parsed.get("current_gpr")
        opex = parsed.get("current_opex")
        if gpr:
            print(f"\n  → Monthly GPR for --analyze  : ${gpr/12:,.0f}")
        if opex:
            print(f"  → Monthly OpEx for --analyze : ${opex/12:,.0f}")
    else:
        print(f"  Units          : {parsed.get('units', 0)}")
        print(f"  Monthly GPR    : {fmt_num(parsed.get('current_gpr'))}")
        mix = parsed.get("unit_mix", [])
        if mix:
            print(f"  Unit Mix:")
            for m in mix:
                mkt = f"  market ${m['market_rent']:,.0f}" if m.get("market_rent") else ""
                print(f"    {m['type']:<8} {m['count']} units @ ${m['current_rent']:,.0f}/mo{mkt}")
    print(f"{'─'*50}\n")


# ─────────────────────────────────────────────
# PDF parsing — OM and Rent Roll
# ─────────────────────────────────────────────

def _require_pdfplumber():
    return _require("pdfplumber", "pip3 install pdfplumber")


def parse_om_pdf(path):
    """
    Parse an Offering Memorandum PDF.
    Returns dict with: asking, units, vintage, entry_cap, exit_cap,
    current_noi_annual, current_gpr_annual, market_gpr_monthly, address, property_name.
    Any value not found is None.
    """
    import re
    pdfplumber = _require_pdfplumber()

    result = {
        "asking": None, "units": None, "vintage": None,
        "entry_cap": None, "exit_cap": None,
        "current_noi_annual": None, "current_gpr_annual": None,
        "market_gpr_monthly": None, "address": None, "property_name": None,
        "source": "OM",
    }

    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    def clean_num(s):
        """Strip $, commas, and spaces from a number string."""
        return float(re.sub(r'[\s,$]', '', s))

    # Asking / list price — value may be on the next line with a space in the number ("$1, 850,000")
    m = re.search(r'(?:List|Asking|Offer(?:ing)?)\s*Price\s*:\s*[\n\s]*\$?\s*([\d][\d,\s]+\d)', text, re.I)
    if m:
        result["asking"] = clean_num(m.group(1))

    # Units
    m = re.search(r'Number\s+of\s+Units\s+(\d+)', text, re.I)
    if not m:
        m = re.search(r'(\d+)\s*[–\-]\s*[Uu]nit', text)
    if m:
        result["units"] = int(m.group(1))

    # Vintage
    m = re.search(r'Year\s+Built\s+(\d{4})', text, re.I)
    if m:
        result["vintage"] = int(m.group(1))

    # Pro-forma / exit cap (check before plain "Cap Rate" to avoid re-matching)
    m = re.search(r'Pro.?[Ff]orma\s*Cap\s*Rate\s*:\s*(\d+\.\d+)%', text, re.I)
    if m:
        result["exit_cap"] = float(m.group(1))

    # Entry cap — skip if preceded by "pro" or "forma" on the same line
    for m in re.finditer(r'Cap\s*Rate\s*:\s*(\d+\.\d+)%', text, re.I):
        preceding = text[max(0, m.start() - 20): m.start()].lower()
        if "pro" not in preceding and "forma" not in preceding:
            result["entry_cap"] = float(m.group(1))
            break

    # Current NOI — first match after "Net Operating Income"
    m = re.search(r'Net\s+Operating\s+Income\s+\$?([\d,]+)', text, re.I)
    if m:
        result["current_noi_annual"] = clean_num(m.group(1))

    # Current GPR — Gross Scheduled Rent
    m = re.search(r'Gross\s+Scheduled\s+Rent\s+\$?([\d,]+)', text, re.I)
    if m:
        result["current_gpr_annual"] = clean_num(m.group(1))

    # Market GPR monthly = avg pro-forma rent × units
    m = re.search(r'Average\s+Pro.?[Ff]orma\s+Rent\s+\$?([\d,]+)', text, re.I)
    if m and result["units"]:
        result["market_gpr_monthly"] = float(m.group(1).replace(',', '')) * result["units"]

    return result


def parse_rent_roll_pdf(path):
    """
    Parse a rent roll PDF (OneSite / typical PM format).
    Returns dict: units, current_gpr (monthly, sum of RESIDENTRENT),
    market_gpr (monthly, sum of market rent column), unit_mix list.
    """
    import re
    pdfplumber = _require_pdfplumber()

    result = {"units": 0, "current_gpr": 0.0, "market_gpr": 0.0,
              "unit_mix": {}, "source": "Rent Roll (PDF)"}

    unit_re    = re.compile(r'^\d{3,5}-[A-Z](?:\s|$)')
    res_rent_re = re.compile(r'RESIDENTRENT\s+([\d,]+\.?\d*)')
    # Market rent: last decimal number on the unit line before the first charge code
    mkt_re     = re.compile(r'(\d+\.\d{2})\s+RESIDENT')

    actual_rents = []
    market_rents = []
    floorplan_rents = {}  # floorplan → [actual rents]

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or '').split('\n')
            current_fp = None
            for line in lines:
                stripped = line.strip()
                if unit_re.match(stripped):
                    # Extract market rent from this unit line
                    m = mkt_re.search(stripped)
                    if m:
                        market_rents.append(float(m.group(1)))
                    # Extract floorplan (second token, e.g. "1x1")
                    parts = stripped.split()
                    current_fp = parts[1] if len(parts) > 1 else None
                elif res_rent_re.search(stripped):
                    m = res_rent_re.search(stripped)
                    if m:
                        amt = float(m.group(1).replace(',', ''))
                        actual_rents.append(amt)
                        if current_fp:
                            floorplan_rents.setdefault(current_fp, []).append(amt)

    result["units"]       = len(actual_rents)
    result["current_gpr"] = sum(actual_rents)
    result["market_gpr"]  = sum(market_rents)

    for fp, rents in floorplan_rents.items():
        result["unit_mix"][fp] = {
            "count": len(rents),
            "avg_actual_rent": round(sum(rents) / len(rents), 0),
        }

    return result


def print_rent_roll_pdf_summary(parsed):
    print(f"\n{'─'*50}")
    print(f"  PARSED: {parsed['source']}")
    print(f"{'─'*50}")
    print(f"  Units            : {parsed['units']}")
    print(f"  Monthly GPR (in-place) : {fmt_num(parsed['current_gpr'])}")
    if parsed.get("market_gpr"):
        print(f"  Monthly GPR (market)   : {fmt_num(parsed['market_gpr'])}")
    mix = parsed.get("unit_mix", {})
    if mix:
        print(f"  Unit Mix:")
        for fp, data in mix.items():
            print(f"    {fp:<8} {data['count']} units @ ${data['avg_actual_rent']:,.0f}/mo avg")
    print(f"{'─'*50}\n")


def print_om_summary(parsed):
    print(f"\n{'─'*50}")
    print(f"  PARSED: {parsed['source']}")
    print(f"{'─'*50}")
    if parsed.get("asking"):
        print(f"  Asking Price    : {fmt_num(parsed['asking'])}")
    if parsed.get("units"):
        print(f"  Units           : {parsed['units']}")
    if parsed.get("vintage"):
        print(f"  Vintage         : {parsed['vintage']}")
    if parsed.get("entry_cap"):
        print(f"  Entry Cap Rate  : {parsed['entry_cap']:.2f}%")
    if parsed.get("exit_cap"):
        print(f"  Pro-Forma Cap   : {parsed['exit_cap']:.2f}%")
    if parsed.get("current_noi_annual"):
        print(f"  Current NOI     : {fmt_num(parsed['current_noi_annual'])}  (annual)")
    if parsed.get("current_gpr_annual"):
        print(f"  Current GPR     : {fmt_num(parsed['current_gpr_annual'])}  (annual)")
    if parsed.get("market_gpr_monthly"):
        print(f"  Market GPR      : {fmt_num(parsed['market_gpr_monthly'])}/mo")
    print(f"{'─'*50}\n")


# ─────────────────────────────────────────────
# Deal Analyzer — populate INPUTS sheet
# ─────────────────────────────────────────────

def populate_analyzer(token, args, metrics=None):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
        sys.exit(1)

    # ── Pull inputs ──
    prop      = getattr(args, 'property', '') or ''
    address   = getattr(args, 'address', '') or prop
    asking    = getattr(args, 'asking', 0) or 0
    offer     = getattr(args, 'offer', None) or asking
    units     = getattr(args, 'units', 0) or 0
    repair    = getattr(args, 'repair', 0) or 0
    ltv       = getattr(args, 'ltv', 0.70)
    down_pct  = round(1 - ltv, 4)
    rate      = getattr(args, 'bridge_rate', 0.0675)
    io_yrs    = getattr(args, 'io_years', 2)
    amort     = getattr(args, 'amort_years', 25)
    closing   = getattr(args, 'closing_costs_pct', 0.06)
    vacancy   = getattr(args, 'vacancy_pct', 0.10)
    other_inc = getattr(args, 'other_income', 0) or 0
    hold_yrs  = getattr(args, 'hold_years', 6)
    exit_cap  = getattr(args, 'exit_cap', None)
    vintage   = getattr(args, 'vintage', None)
    unit_mix  = getattr(args, 'unit_mix', None)
    curr_gpr  = getattr(args, 'current_gpr', None)   # monthly
    mkt_gpr   = getattr(args, 'market_gpr', None)    # monthly

    print(f"Downloading Deal Analyzer template ({DEAL_ANALYZER_ID})...")
    local_path = "/tmp/deal_analyzer_source.xlsx"
    fetch_doc(token, DEAL_ANALYZER_ID, local_path)

    wb = openpyxl.load_workbook(local_path)
    ws = wb["INPUTS"]

    # ── INPUTS sheet ──
    # Deal identity
    ws["A2"] = address or prop

    # Sources & uses
    if asking:  ws["B4"] = asking
    if offer:   ws["D4"] = offer
    ws["C5"] = down_pct          # down payment % (spreadsheet: 1 - LTV)
    if units:   ws["D6"] = units
    ws["D8"] = repair            # 0 if no capex — always write to clear template value
    ws["D9"] = "no"              # reno in loan — can be overridden manually
    if vintage: ws["D12"] = vintage

    # Financing
    ws["F14"] = rate
    ws["H12"] = io_yrs           # I/O years (SENIOR LOAN reads this via F10=H12*12)
    ws["C16"] = closing          # closing costs %

    # Vacancy (current and proforma)
    ws["C39"] = vacancy
    ws["F39"] = vacancy

    # Other income (annual)
    if other_inc:
        ws["D40"] = other_inc
        ws["G40"] = other_inc

    # Hold period + exit parameters (column R)
    ws["R35"] = hold_yrs
    if exit_cap is not None:
        ws["R37"] = exit_cap / 100   # store as decimal (e.g. 6.0 → 0.06)
    ws["R38"] = 0.05                 # selling costs % (5% default)

    # ── SENIOR LOAN sheet — amortization ──
    ws_loan = wb["SENIOR LOAN"]
    ws_loan["C9"] = amort

    # ── Cell comments (source attribution on every input cell) ──
    try:
        from openpyxl.comments import Comment as _Comment
        def _note(cell, text):
            ws[cell].comment = _Comment(text, "Olive Tree AIOS")

        _note("A2",  f"Property: {prop or address} | Analyzed: {TODAY_ISO}")
        _note("B4",  "Asking price — from OM PDF. Adjust D4 to test different offer prices.")
        _note("D4",  "Offer price — change this cell to model different offer scenarios.")
        _note("C5",  f"Down payment {down_pct:.0%} (LTV {ltv:.0%}). Adjust to test higher/lower leverage.")
        _note("D6",  f"{units} total units per OM. Rent roll shows {units} in unit mix.")
        _note("D8",  f"Capex/repair budget — ${repair:,.0f}. Add estimate if value-add work required.")
        _note("D9",  "Set to 'no'. Change to 'yes' if renovation is included in the loan.")
        _note("D12", f"Built {vintage or 'unknown'}. Older builds: higher capex reserve; check mechanical systems.")
        _note("F14", f"Bridge loan rate {rate:.2%}. Update if lender quotes a different rate.")
        _note("H12", f"{io_yrs}-year interest-only period. P+I begins year {io_yrs + 1}.")
        _note("C16", f"Closing costs {closing:.0%} of offer. Includes lender, title, and legal fees.")
        _note("C39", f"Current vacancy {vacancy:.0%}. Based on rent roll / market assumption.")
        _note("F39", f"Proforma vacancy {vacancy:.0%}. Adjust down if market is tight.")
        _note("R35", f"{hold_yrs}-year hold period. Sale modeled at year {hold_yrs}.")
        if exit_cap is not None:
            _note("R37", f"Exit cap {exit_cap:.2f}% from OM pro-forma. May expand if market softens — model conservatively.")
        _note("R38", "Selling costs 5% (broker + closing costs at exit).")

        # Unit mix comments
        if curr_gpr and units:
            _note("C21", f"T-12 derived: ${curr_gpr*12:,.0f}/yr ÷ {units} units ≈ ${curr_gpr/units:,.0f}/unit/mo")
        if mkt_gpr and units:
            _note("F21", f"OM market GPR: ${mkt_gpr:,.0f}/mo ÷ {units} units = ${mkt_gpr/units:,.0f}/unit/mo")

        # NOTES section — deal analysis summary
        if metrics:
            irr  = metrics.get("irr_estimate")
            em   = metrics.get("equity_multiple")
            dscr = metrics.get("dscr")
            coc3 = metrics.get("coc_yr3")
            note_lines = [
                f"OLIVE TREE AIOS — DEAL ANALYSIS  ({TODAY_ISO})",
                f"Property  : {prop or address}",
                f"Asking    : ${asking:,.0f}  |  Units: {units}  |  PPU: ${asking/units:,.0f}" if units else f"Asking: ${asking:,.0f}",
                "",
                f"IRR (est.)       : {irr:.1%}" if irr else "IRR (est.)       : N/A",
                f"Equity Multiple  : {em:.2f}x" if em else "Equity Multiple  : N/A",
                f"DSCR             : {dscr:.2f}x" if dscr else "DSCR             : N/A",
                f"CoC Yr3          : {coc3:.1%}" if coc3 else "CoC Yr3          : N/A",
            ]
            rec  = metrics.get("_rec", "")
            if rec:
                note_lines += ["", f"RECOMMENDATION: {rec}"]
            ws["A58"] = "\n".join(note_lines)
    except Exception:
        pass  # comments are non-critical — never block the upload

    # ── Unit mix (rows 21-34) ──
    # Col A=count, B=type name, C=current rent/unit, F=proforma rent/unit
    mix = None
    if unit_mix:
        try:
            mix = json.loads(unit_mix) if isinstance(unit_mix, str) else unit_mix
        except (json.JSONDecodeError, TypeError):
            pass

    if mix:
        total_mix_units = sum(u.get("count", 0) for u in mix)
        for i, unit in enumerate(mix[:14]):
            rn     = 21 + i
            count  = unit.get("count", 0) or 0
            cur_r  = unit.get("current_rent", 0) or 0
            mkt_r  = unit.get("market_rent", None)
            if mkt_r is None and mkt_gpr and total_mix_units:
                mkt_r = round(mkt_gpr / total_mix_units, 0)
            ws.cell(row=rn, column=1, value=count)
            ws.cell(row=rn, column=2, value=unit.get("type", ""))
            ws.cell(row=rn, column=3, value=cur_r)
            ws.cell(row=rn, column=6, value=mkt_r or 0)
    elif units:
        # No unit mix — single row with avg per unit
        cur_per = round(curr_gpr / units, 0) if curr_gpr and units else None
        mkt_per = round(mkt_gpr  / units, 0) if mkt_gpr  and units else None
        ws["A21"] = units
        ws["B21"] = "Units"
        if cur_per: ws["C21"] = cur_per
        if mkt_per: ws["F21"] = mkt_per

    # ── Save + upload as native Google Sheet ──
    # Uploading with mimeType=spreadsheet converts xlsx → Google Sheet,
    # so all formulas calculate live and the file is directly editable.
    sheet_name = f"{address or prop} - Deal Analyzer 0-50"
    out_path   = f"/tmp/{TODAY_ISO}_deal_analyzer_output.xlsx"
    try:
        wb.save(out_path)
    finally:
        wb.close()
    print(f"\nUploading as Google Sheet: '{sheet_name}'...")

    with open(out_path, "rb") as f:
        file_bytes = f.read()

    metadata = json.dumps({
        "name": sheet_name,
        "mimeType": "application/vnd.google-apps.spreadsheet"
    })
    boundary = "boundary_olive_tree_analyzer"
    body = (
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n"
        f"{metadata}\r\n"
        f"--{boundary}\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_bytes + f"\r\n--{boundary}--".encode()

    r = requests.post(
        f"{UPLOAD_BASE}?uploadType=multipart",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
        data=body,
        timeout=120,
    )
    r.raise_for_status()
    file_id = r.json().get("id")
    print(f"✅ Created: {sheet_name}")
    print(f"   Drive ID : {file_id}")
    print(f"   Open     : https://docs.google.com/spreadsheets/d/{file_id}/edit")
    return file_id


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def build_parser():
    p = argparse.ArgumentParser(description="Olive Tree — Deal Analysis")

    # Mode flags
    p.add_argument("--analyze",            action="store_true", help="Run underwriting analysis")
    p.add_argument("--log-deal",           action="store_true", help="Log deal to Deal Sourcing sheet")
    p.add_argument("--populate-analyzer",  action="store_true", help="Populate Deal Analyzer xlsx and upload")
    p.add_argument("--fetch-docs",         action="store_true", help="Download doc from Drive")
    p.add_argument("--dry-run",            action="store_true", help="Print only — no writes")
    p.add_argument("--yes",                action="store_true", help="Auto-confirm buy-box override prompt")

    # Deal identifiers
    p.add_argument("--property",      type=str, help="Property name")
    p.add_argument("--address",       type=str, help="Property address")
    p.add_argument("--market",        type=str, help="Market name (e.g. 'Chamblee, GA')")
    p.add_argument("--zip",           type=str, help="Zip code")

    # Deal financials
    p.add_argument("--asking",        type=float, help="Asking price")
    p.add_argument("--units",         type=int,   help="Number of units")
    p.add_argument("--offer",              type=float, help="Offer price (defaults to asking if not set)")
    p.add_argument("--repair",             type=float, help="Repair / capex budget")
    p.add_argument("--io-years",           type=int,   default=2,    help="Interest-only period in years (default: 2)")
    p.add_argument("--amort-years",        type=int,   default=25,   help="Amortization years (default: 25)")
    p.add_argument("--rent-growth",        type=float, default=0.03, help="Annual rent growth after yr2 (default: 0.03)")
    p.add_argument("--expense-growth",     type=float, default=0.02, help="Annual expense growth (default: 0.02)")
    p.add_argument("--other-income",       type=float, default=0,    help="Other annual income (laundry, parking, etc.)")
    p.add_argument("--closing-costs-pct",  type=float, default=0.06, help="Closing costs as %% of offer (default: 0.06)")
    p.add_argument("--current-gpr",   type=float, help="Current gross potential rent (monthly total)")
    p.add_argument("--current-opex",  type=float, help="Current operating expenses (monthly total)")
    p.add_argument("--market-gpr",    type=float, help="Stabilized market GPR (monthly total)")
    p.add_argument("--vacancy-pct",   type=float, default=0.10, help="Vacancy % as decimal (default: 0.10)")

    # Rentometer market rent inputs (auto-lookup if --beds + --address provided)
    p.add_argument("--beds",      type=int, choices=[1, 2, 3, 4],
                   help="Dominant bedroom type for Rentometer lookup (1–4)")
    p.add_argument("--baths",     type=str, choices=["1", "1.5+"],
                   help="Bath filter for Rentometer: '1' or '1.5+' (optional)")
    p.add_argument("--om-rent",   type=float,
                   help="OM asking rent per unit/mo — compared against Rentometer comps")

    p.add_argument("--bridge-rate",   type=float, default=0.0675, help="Bridge loan rate (default: 0.0675)")
    p.add_argument("--ltv",           type=float, default=0.70,  help="Loan-to-value (default: 0.70)")
    p.add_argument("--hold-years",    type=int,   default=6,     help="Hold period in years (default: 6)")
    p.add_argument("--entry-cap",     type=float, help="Entry cap rate (as number, e.g. 5.5 = 5.5%%)")
    p.add_argument("--exit-cap",      type=float, help="Exit cap rate (as number, e.g. 6.0 = 6.0%%)")
    p.add_argument("--vintage",       type=int,   help="Building vintage year")

    # Sheet logging
    p.add_argument("--stage",         type=str, default="New",  help="Deal stage for sheet")
    p.add_argument("--platform",      type=str, default="",     help="Source platform (Crexi/LoopNet/Email)")
    p.add_argument("--brokerage",     type=str, default="",     help="Brokerage name")
    p.add_argument("--broker-name",   type=str, default="",     help="Broker name")
    p.add_argument("--broker-email",  type=str, default="",     help="Broker email")
    p.add_argument("--broker-phone",  type=str, default="",     help="Broker phone")
    p.add_argument("--notes",         type=str, default="",     help="Notes for sheet")

    # Deal Analyzer
    p.add_argument("--unit-mix",      type=str, help='JSON unit mix: [{"type":"1BR","count":10,"current_rent":800,"market_rent":950}]')

    # Archive
    p.add_argument("--archive",       action="store_true",
                   help="On PASS verdict: create dated Drive folder and upload Deal Summary")
    p.add_argument("--archive-files", nargs="*", default=[], metavar="FILE",
                   help="Local files to upload to the archive folder (OM, T-12, Rent Roll, etc.)")

    # Doc fetch
    p.add_argument("--drive-id",      type=str, help="Google Drive file ID to download")
    p.add_argument("--gmail-id",      type=str, help="Gmail message ID to fetch attachments from")

    # Excel / PDF parsing
    p.add_argument("--parse-excel",   type=str, metavar="FILE",
                   help="Path to rent roll or T-12 xlsx. Parsed values feed --analyze automatically.")
    p.add_argument("--excel-type",    type=str, choices=["auto", "rent-roll", "t12"], default="auto",
                   help="Excel file type (default: auto-detect)")
    p.add_argument("--om",            type=str, metavar="FILE",
                   help="Path to Offering Memorandum PDF. Extracts asking, units, cap rates, NOI.")
    p.add_argument("--rent-roll",     type=str, metavar="FILE",
                   help="Path to rent roll PDF. Extracts in-place and market rents.")

    return p


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not any([args.analyze, args.log_deal, args.populate_analyzer, args.fetch_docs,
                getattr(args, 'parse_excel', None),
                getattr(args, 'om', None),
                getattr(args, 'rent_roll', None)]):
        parser.print_help()
        sys.exit(0)

    # Snapshot explicit CLI values before any parsing — used to enforce priority later
    _cli_asking    = args.asking
    _cli_units     = args.units
    _cli_curr_gpr  = args.current_gpr
    _cli_curr_opex = args.current_opex
    _cli_mkt_gpr   = args.market_gpr
    _cli_entry_cap = args.entry_cap
    _cli_exit_cap  = args.exit_cap
    _cli_vintage   = args.vintage

    # ── --om: parse OM PDF (lowest priority — overridden by rent roll and T-12) ──
    if getattr(args, 'om', None):
        om = parse_om_pdf(args.om)
        print_om_summary(om)
        if om.get("asking")    and not _cli_asking:     args.asking     = om["asking"]
        if om.get("units")     and not _cli_units:      args.units      = om["units"]
        if om.get("vintage")   and not _cli_vintage:    args.vintage    = om["vintage"]
        if om.get("entry_cap") and not _cli_entry_cap:  args.entry_cap  = om["entry_cap"]
        if om.get("exit_cap")  and not _cli_exit_cap:   args.exit_cap   = om["exit_cap"]
        if om.get("market_gpr_monthly") and not _cli_mkt_gpr:
            args.market_gpr = om["market_gpr_monthly"]
        if om.get("current_gpr_annual") and not _cli_curr_gpr:
            args.current_gpr  = om["current_gpr_annual"] / 12
            if om.get("current_noi_annual") and not _cli_curr_opex:
                args.current_opex = (om["current_gpr_annual"] - om["current_noi_annual"]) / 12

    # ── --rent-roll: parse rent roll PDF (fills gaps not covered by OM) ──
    # Note: rent roll units = occupied count; OM has total units — OM takes priority.
    # Note: rent roll market GPR covers only occupied units; OM proforma GPR covers all — OM takes priority.
    # T-12 will later override OM-derived current_gpr/opex regardless.
    if getattr(args, 'rent_roll', None):
        rr = parse_rent_roll_pdf(args.rent_roll)
        print_rent_roll_pdf_summary(rr)
        if rr.get("units")       and not args.units:       args.units       = rr["units"]
        if rr.get("current_gpr") and not args.current_gpr: args.current_gpr = rr["current_gpr"]
        if rr.get("market_gpr")  and not args.market_gpr:  args.market_gpr  = rr["market_gpr"]

    # Validate required fields
    if args.analyze or args.log_deal or args.populate_analyzer:
        if not args.asking:
            print("ERROR: --asking is required (or provide --om to parse from OM PDF)")
            sys.exit(1)
        if not args.units:
            print("ERROR: --units is required (or provide --om / --rent-roll to parse)")
            sys.exit(1)

    # Buy box check
    zip_s = str(args.zip or '')
    if zip_s and zip_s not in BUY_BOX:
        print(f"⚠️  ZIP {zip_s} is outside the active buy box.")
        print(f"   Active zips: {', '.join(BUY_BOX.keys())}")
        if getattr(args, 'yes', False):
            print("   Continuing (--yes flag set).")
        else:
            try:
                confirm = input("   Continue anyway? (y/n): ").strip().lower()
            except EOFError:
                confirm = 'n'
            if confirm != 'y':
                print("Stopped.")
                sys.exit(0)

    token = None
    if not args.dry_run:
        try:
            token = get_token()
        except Exception as e:
            print(f"ERROR: Auth failed — {e}")
            print("Run: gws auth login -s gmail,sheets,drive")
            sys.exit(1)

    # ── --fetch-docs ──
    if args.fetch_docs:
        if not args.drive_id:
            print("ERROR: --drive-id required with --fetch-docs")
            sys.exit(1)
        fetch_doc(token, args.drive_id)
        return

    # ── --gmail-id: download deal attachments, classify, feed into analysis ──
    if args.gmail_id:
        docs = fetch_gmail_attachments(token, args.gmail_id)
        if not any([docs["t12"], docs["rent-roll"], docs["om"]]):
            print("ERROR: No recognizable deal documents (xlsx/pdf) found in that message.")
            sys.exit(1)

        print(f"\nAttachments classified:")
        print(f"  T-12       : {docs['t12']      or '—'}")
        print(f"  Rent Roll  : {docs['rent-roll'] or '—'}")
        print(f"  OM (PDF)   : {docs['om']        or '—'}")
        if docs["other"]:
            print(f"  Other      : {', '.join(docs['other'])}")
        print()

        if docs["t12"] and not getattr(args, "parse_excel", None):
            args.parse_excel = docs["t12"]
            args.excel_type  = "auto"
        if docs["rent-roll"] and not getattr(args, "rent_roll", None):
            rr_ext = docs["rent-roll"].rsplit(".", 1)[-1].lower()
            if rr_ext in ("xlsx", "xls"):
                # xlsx rent roll — feed through --parse-excel only if no T-12 already set
                if not getattr(args, "parse_excel", None):
                    args.parse_excel = docs["rent-roll"]
                    args.excel_type  = "rent-roll"
            else:
                args.rent_roll = docs["rent-roll"]
        if docs["om"] and not getattr(args, "om", None):
            args.om = docs["om"]

    # ── --parse-excel (standalone or feed into --analyze) ──
    if getattr(args, 'parse_excel', None):
        parsed = parse_excel_file(args.parse_excel, args.excel_type)
        print_parse_summary(parsed)

        if args.analyze:
            # T-12 has highest priority after explicit CLI — overrides OM and rent roll
            src = parsed.get("source", "")
            if src == "T-12":
                gpr_annual  = parsed.get("current_gpr")
                opex_annual = parsed.get("current_opex")
                if gpr_annual  and not _cli_curr_gpr:   args.current_gpr  = gpr_annual  / 12
                if opex_annual and not _cli_curr_opex:  args.current_opex = opex_annual / 12
            else:  # xlsx rent roll (fills gaps not covered by OM — same priority as PDF rent roll)
                if parsed.get("units")       and not args.units:      args.units       = parsed["units"]
                if parsed.get("current_gpr") and not args.current_gpr: args.current_gpr = parsed["current_gpr"]
                mix = parsed.get("unit_mix", [])
                if mix and not getattr(args, 'unit_mix', None):
                    args.unit_mix = json.dumps(mix)

    # ── --analyze ──
    if args.analyze:
        # Market rent lookup via Rentometer
        # Runs when --beds is provided and --market-gpr is not already set.
        # Uses median per-unit rent * units as the stabilized GPR input.
        _beds = getattr(args, 'beds', None)
        _om_rent = getattr(args, 'om_rent', None)
        if _beds and args.address and not args.market_gpr and os.getenv("RENTOMETER_API_KEY", "").strip():
            try:
                sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
                import rentometer as _rentometer
                _baths = getattr(args, 'baths', None)
                print(f"\nPulling Rentometer comps for {_beds}BR at {args.address}...")
                _rm = _rentometer.lookup(
                    address=args.address,
                    beds=_beds,
                    baths=_baths,
                    om_rent=_om_rent,
                )
                _rentometer.print_report(_rm, _beds, _baths)
                if _rm and _rm.get("median") and args.units:
                    # Set market_gpr = median_per_unit * units (total monthly market GPR)
                    args.market_gpr = _rm["median"] * args.units
                    print(f"  Using Rentometer median ${_rm['median']:,.0f}/unit × {args.units} units"
                          f" = ${args.market_gpr:,.0f}/mo as market GPR\n")
            except Exception as _e:
                print(f"  ⚠️  Rentometer lookup skipped: {_e}\n")
        elif _beds and args.address and not args.market_gpr:
            print("  ℹ️  RENTOMETER_API_KEY not set — skipping market rent lookup.\n"
                  "     Add key to .env to enable auto comps.\n")

        metrics = calculate_metrics(args)
        rec, passes, fails, warnings = score_deal(metrics, zip_s)
        print_analysis(args, metrics, rec, passes, fails, warnings)

        prop    = args.property or ''
        address = args.address or ''

        if not args.dry_run and token:
            # Deduplicate
            if prop and deal_exists(token, prop, address):
                print(f"ℹ️  Deal '{prop}' already in sheet — skipping log.")
            else:
                # Auto-log on analyze (stage = analyzing or pass)
                stage_val = "Pass" if rec == "PASS" else "Analyzing"
                rec_label = {"PURSUE_LOI": "PURSUE LOI", "MORE_INFO": "MORE INFO NEEDED", "PASS": "PASS"}[rec]
                notes_val = args.notes or f"{rec_label}"
                args.stage = stage_val
                args.notes = notes_val
                log_deal(token, args, metrics)

            # Archive on PASS if --archive flag is set
            if rec == "PASS" and getattr(args, 'archive', False):
                if not address:
                    print("⚠️  --archive requires --address to name the folder. Skipping.")
                else:
                    import sys as _sys
                    import os as _os
                    _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
                    import deal_archive
                    deal_archive.archive_deal(
                        token=token,
                        address=address,
                        property_name=prop,
                        metrics=metrics,
                        notes=args.notes,
                        files=getattr(args, 'archive_files', []),
                    )

        # Populate Deal Analyzer if --populate-analyzer also set (works with --dry-run too)
        if args.populate_analyzer:
            sheet_name = f"{address or prop} - Deal Analyzer 0-50"
            if args.dry_run:
                print(f"[DRY RUN] Would create Google Sheet: '{sheet_name}'")
            else:
                metrics["_rec"] = {"PURSUE_LOI": "PURSUE LOI ✅", "MORE_INFO": "MORE INFO NEEDED ⚠️", "PASS": "PASS ❌"}.get(rec, rec)
                populate_analyzer(token, args, metrics=metrics)

    # ── --log-deal ──
    elif args.log_deal:
        if args.dry_run:
            print(f"[DRY RUN] Would log: {args.property} to Deal Sourcing tab")
            return
        log_deal(token, args)

    # ── --populate-analyzer (standalone, without --analyze) ──
    elif args.populate_analyzer:
        if args.dry_run:
            sheet_name = f"{getattr(args,'address','') or args.property} - Deal Analyzer 0-50"
            print(f"[DRY RUN] Would create Google Sheet: '{sheet_name}'")
            return
        populate_analyzer(token, args)


if __name__ == "__main__":
    main()
