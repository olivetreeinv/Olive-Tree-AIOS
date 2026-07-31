"""
MFS Deal Desk — document parsers (OM PDF, rent-roll PDF/XLSX/CSV, T-12 XLSX/CSV).

Vendored from Olive Tree AIOS `scripts/deal_analysis.py` (2026-07-28).
Deterministic-first: these regex/keyword parsers run before any LLM extraction.
openpyxl-only (no pandas). CSV support added via the shared row-scan helpers.
"""

import csv
import re

# T-12 parsing keywords
_T12_INCOME_KEYS  = ["gross potential rent", "gross potential income", "scheduled base rent",
                     "gross rents", "rental income", "total income", "total revenue",
                     "effective gross income"]
_T12_EXPENSE_KEYS = ["total expenses", "total operating expenses", "operating expenses", "total opex"]
_T12_NOI_KEYS     = ["net operating income", "noi"]
_T12_VACANCY_KEYS = ["vacancy", "vacancy loss", "credit loss"]

# Rent roll parsing keywords
_RR_RENT_COLS   = ["rent", "current rent", "actual rent", "monthly rent", "contract rent", "lease rent"]
_RR_MARKET_COLS = ["market rent", "market", "asking rent"]
_RR_BED_COLS    = ["bed", "br", "bedroom", "unit type", "type", "floorplan"]
_RR_UNIT_COLS   = ["unit", "apt", "apartment", "unit no", "unit #", "unit#"]
_RR_ANCHOR_COLS = _RR_UNIT_COLS + _RR_BED_COLS + _RR_RENT_COLS + ["tenant", "status", "occupied"]

# Excel/CSV type detection signals
_RR_SIGNALS  = ["tenant", "unit no", "apt", "lease", "move-in"]
_T12_SIGNALS = ["operating expenses", "net operating", "vacancy loss", "gross potential"]


# ─────────────────────────────────────────────
# Row-level scanners (shared by xlsx and csv front-ends)
# ─────────────────────────────────────────────

def detect_type_rows(rows):
    """'rent-roll' or 't12' from the first 30 rows of cell values."""
    parts = []
    for i, row in enumerate(rows):
        if i >= 30:
            break
        parts.extend(str(v).lower() for v in row if v is not None and str(v).strip())
    text = " ".join(parts)
    rr  = sum(w in text for w in _RR_SIGNALS)
    t12 = sum(w in text for w in _T12_SIGNALS)
    return "rent-roll" if rr >= t12 else "t12"


def _scan_t12_rows(rows, result):
    """Keyword-match income/expense/NOI/vacancy lines across rows."""
    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0]).lower().strip()
        nums = []
        for v in row[1:]:
            try:
                f = float(v)
                if f > 0:
                    nums.append(f)
            except (TypeError, ValueError):
                pass
        if not nums:
            continue
        median_val = sorted(nums)[len(nums) // 2]
        annual = nums[-1] if (len(nums) > 1 and nums[-1] >= median_val * 8) else sum(nums[:12])

        if any(label.startswith(k) for k in _T12_INCOME_KEYS) and result["current_gpr"] is None:
            result["current_gpr"] = annual
        elif any(label.startswith(k) for k in _T12_EXPENSE_KEYS) and result["current_opex"] is None:
            result["current_opex"] = annual
        elif any(label.startswith(k) for k in _T12_NOI_KEYS) and result["current_noi"] is None:
            result["current_noi"] = annual
        elif any(label.startswith(k) for k in _T12_VACANCY_KEYS) and result["vacancy_annual"] is None:
            result["vacancy_annual"] = annual
    return result


def _finish_t12(result):
    if result["current_noi"] and result["current_gpr"] and result["current_opex"] is None:
        result["current_opex"] = result["current_gpr"] - result["current_noi"]
    if result["current_noi"] and result["current_opex"] and result["current_gpr"] is None:
        result["current_gpr"] = result["current_noi"] + result["current_opex"]
    return result


def _scan_rent_roll_rows(row_iter, result):
    """Find the header row, then sum rents / build unit mix."""
    headers = None
    for row in row_iter:
        row_text = [str(v).lower().strip() if v is not None else "" for v in row]
        if sum(1 for v in row_text if any(k in v for k in _RR_ANCHOR_COLS)) >= 2:
            headers = row_text
            break
    if headers is None:
        return None

    rent_col   = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_RENT_COLS)), None)
    market_col = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_MARKET_COLS)), None)
    bed_col    = next((i for i, h in enumerate(headers) if any(k in h for k in _RR_BED_COLS)), None)
    if rent_col is None:
        return None

    unit_rents, unit_markets = {}, {}
    total_gpr, unit_count = 0.0, 0
    for row in row_iter:
        if not row or len(row) <= rent_col or row[rent_col] is None:
            continue
        try:
            rent_val = float(row[rent_col])
        except (ValueError, TypeError):
            continue
        if rent_val <= 0:
            continue
        total_gpr += rent_val
        unit_count += 1
        if bed_col is not None and len(row) > bed_col and row[bed_col] is not None:
            bed = str(row[bed_col]).strip()
            unit_rents.setdefault(bed, []).append(rent_val)
            if market_col is not None and len(row) > market_col and row[market_col] is not None:
                try:
                    mkt = float(row[market_col])
                    if mkt > 0:
                        unit_markets.setdefault(bed, []).append(mkt)
                except (ValueError, TypeError):
                    pass

    if unit_count == 0:
        return None
    result["units"] = unit_count
    result["current_gpr"] = total_gpr
    for bed, rents in unit_rents.items():
        entry = {"type": bed, "count": len(rents),
                 "current_rent": round(sum(rents) / len(rents), 0)}
        mkts = unit_markets.get(bed, [])
        if mkts:
            entry["market_rent"] = round(sum(mkts) / len(mkts), 0)
        result["unit_mix"].append(entry)
    return result


# ─────────────────────────────────────────────
# XLSX front-ends (openpyxl)
# ─────────────────────────────────────────────

def detect_excel_type(path):
    """Detect 'rent-roll' or 't12' by scanning first 30 rows."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return detect_type_rows(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def parse_t12_xlsx(path):
    """Returns dict: current_gpr, current_opex, current_noi, vacancy_annual (annual $)."""
    import openpyxl
    result = {"current_gpr": None, "current_opex": None, "current_noi": None,
              "vacancy_annual": None, "source": "T-12"}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _scan_t12_rows(ws.iter_rows(values_only=True), result)
            if result["current_gpr"] or result["current_noi"]:
                break
    finally:
        wb.close()
    return _finish_t12(result)


def parse_rent_roll_xlsx(path):
    """Returns dict: units, current_gpr (monthly total), unit_mix list."""
    import openpyxl
    result = {"units": 0, "current_gpr": 0.0, "unit_mix": [], "source": "Rent Roll"}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if _scan_rent_roll_rows(ws.iter_rows(values_only=True), result):
                break
    finally:
        wb.close()
    return result


# ─────────────────────────────────────────────
# CSV front-ends
# ─────────────────────────────────────────────

def _csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        yield from csv.reader(f)


def detect_csv_type(path):
    return detect_type_rows(_csv_rows(path))


def parse_t12_csv(path):
    result = {"current_gpr": None, "current_opex": None, "current_noi": None,
              "vacancy_annual": None, "source": "T-12"}
    _scan_t12_rows(_csv_rows(path), result)
    return _finish_t12(result)


def parse_rent_roll_csv(path):
    result = {"units": 0, "current_gpr": 0.0, "unit_mix": [], "source": "Rent Roll"}
    _scan_rent_roll_rows(_csv_rows(path), result)
    return result


def parse_tabular_file(path, doc_type="auto"):
    """Dispatch xlsx/csv → the right parser. doc_type: auto | rent-roll | t12."""
    is_csv = str(path).lower().endswith(".csv")
    if doc_type == "auto":
        doc_type = detect_csv_type(path) if is_csv else detect_excel_type(path)
    if doc_type == "rent-roll":
        return parse_rent_roll_csv(path) if is_csv else parse_rent_roll_xlsx(path)
    return parse_t12_csv(path) if is_csv else parse_t12_xlsx(path)


# ─────────────────────────────────────────────
# PDF parsers (pdfplumber)
# ─────────────────────────────────────────────

def parse_om_pdf(path):
    """Parse an Offering Memorandum PDF (regex, tuned to common layouts).
    Any value not found is None — the LLM extraction fallback fills gaps."""
    import pdfplumber

    result = {
        "asking": None, "units": None, "vintage": None,
        "entry_cap": None, "exit_cap": None,
        "current_noi_annual": None, "current_gpr_annual": None,
        "market_gpr_monthly": None, "address": None, "property_name": None,
        "source": "OM",
    }

    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    def clean_num(s):
        return float(re.sub(r'[\s,$]', '', s))

    m = re.search(r'(?:List|Asking|Offer(?:ing)?)\s*Price\s*:\s*[\n\s]*\$?\s*([\d][\d,\s]+\d)', text, re.I)
    if m:
        result["asking"] = clean_num(m.group(1))

    m = re.search(r'Number\s+of\s+Units\s+(\d+)', text, re.I)
    if not m:
        m = re.search(r'(\d+)\s*[–\-]\s*[Uu]nit', text)
    if m:
        result["units"] = int(m.group(1))

    m = re.search(r'Year\s+Built\s+(\d{4})', text, re.I)
    if m:
        result["vintage"] = int(m.group(1))

    m = re.search(r'Pro.?[Ff]orma\s*Cap\s*Rate\s*:\s*(\d+\.\d+)%', text, re.I)
    if m:
        result["exit_cap"] = float(m.group(1))

    for m in re.finditer(r'Cap\s*Rate\s*:\s*(\d+\.\d+)%', text, re.I):
        preceding = text[max(0, m.start() - 20): m.start()].lower()
        if "pro" not in preceding and "forma" not in preceding:
            result["entry_cap"] = float(m.group(1))
            break

    m = re.search(r'Net\s+Operating\s+Income\s+\$?([\d,]+)', text, re.I)
    if m:
        result["current_noi_annual"] = clean_num(m.group(1))

    m = re.search(r'Gross\s+Scheduled\s+Rent\s+\$?([\d,]+)', text, re.I)
    if m:
        result["current_gpr_annual"] = clean_num(m.group(1))

    m = re.search(r'Average\s+Pro.?[Ff]orma\s+Rent\s+\$?([\d,]+)', text, re.I)
    if m and result["units"]:
        result["market_gpr_monthly"] = float(m.group(1).replace(',', '')) * result["units"]

    return result


def parse_rent_roll_pdf(path):
    """Parse a rent roll PDF (OneSite / typical PM export format).
    Returns dict: units, current_gpr (monthly), market_gpr (monthly), unit_mix."""
    import pdfplumber

    result = {"units": 0, "current_gpr": 0.0, "market_gpr": 0.0,
              "unit_mix": {}, "source": "Rent Roll (PDF)"}

    unit_re     = re.compile(r'^\d{3,5}-[A-Z](?:\s|$)')
    res_rent_re = re.compile(r'RESIDENTRENT\s+([\d,]+\.?\d*)')
    mkt_re      = re.compile(r'(\d+\.\d{2})\s+RESIDENT')

    actual_rents, market_rents = [], []
    floorplan_rents = {}

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or '').split('\n')
            current_fp = None
            for line in lines:
                stripped = line.strip()
                if unit_re.match(stripped):
                    m = mkt_re.search(stripped)
                    if m:
                        market_rents.append(float(m.group(1)))
                    parts = stripped.split()
                    current_fp = parts[1] if len(parts) > 1 else None
                elif res_rent_re.search(stripped):
                    m = res_rent_re.search(stripped)
                    if m:
                        amt = float(m.group(1).replace(',', ''))
                        actual_rents.append(amt)
                        if current_fp:
                            floorplan_rents.setdefault(current_fp, []).append(amt)

    result["units"]       = len(actual_rents)
    result["current_gpr"] = sum(actual_rents)
    result["market_gpr"]  = sum(market_rents)
    for fp, rents in floorplan_rents.items():
        result["unit_mix"][fp] = {
            "count": len(rents),
            "avg_actual_rent": round(sum(rents) / len(rents), 0),
        }
    return result


# ─────────────────────────────────────────────
# First-pass classification (no LLM)
# ─────────────────────────────────────────────

def classify_filename(name):
    """Best-effort doc-type guess from a filename. Returns om|t12|rent_roll|None."""
    n = name.lower()
    if any(k in n for k in ("t-12", "t12", "trailing", "operating statement", "income statement", "p&l", "pnl")):
        return "t12"
    if any(k in n for k in ("rent roll", "rent_roll", "rentroll", "rr_")):
        return "rent_roll"
    if any(k in n for k in ("om", "offering", "memorandum", "brochure", "package")):
        return "om"
    return None


def pdf_first_page_text(path, max_chars=4000):
    """First-page text for cheap keyword/LLM classification."""
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        if not pdf.pages:
            return ""
        return (pdf.pages[0].extract_text() or "")[:max_chars]
